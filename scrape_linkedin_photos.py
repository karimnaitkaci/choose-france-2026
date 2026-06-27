#!/usr/bin/env python3
"""
Script pour récupérer les photos de profil LinkedIn de tous les contacts
Utilise Playwright + authentification
"""

import json
import time
import sys
from pathlib import Path
from datetime import datetime

try:
    from playwright.sync_api import sync_playwright
    import openpyxl
except ImportError:
    print("❌ Les dépendances manquent.")
    print("Installe d'abord:")
    print("  pip install playwright openpyxl")
    print("\nPuis lance:")
    print("  playwright install")
    sys.exit(1)

# Configuration
EXCEL_FILE = "base_recrutement_luxe_france.xlsx"
OUTPUT_FILE = "contacts_with_photos.json"
EMAIL = "karimnait@hotmail.fr"
DELAY_BETWEEN_PROFILES = 3  # secondes entre chaque profil

def load_contacts():
    """Charge les contacts du fichier Excel"""
    print(f"📂 Chargement de {EXCEL_FILE}...")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Tous les Contacts']

    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    contacts = []
    for row in range(2, ws.max_row + 1):
        contact = {}
        for col, header in enumerate(headers, 1):
            cell_value = ws.cell(row, col).value
            contact[header] = cell_value if cell_value is not None else ""
        contacts.append(contact)

    print(f"✓ {len(contacts)} contacts chargés\n")
    return contacts

def extract_photo_url(page, linkedin_url):
    """Extrait l'URL de la photo de profil"""
    try:
        # Accéder au profil
        page.goto(linkedin_url, wait_until="networkidle", timeout=15000)
        time.sleep(1)

        # Chercher l'image de profil (plusieurs sélecteurs possibles)
        selectors = [
            'img[alt*="profile"][alt*="photo"]',
            'img[data-delayed-url]',
            'img.ProfilePhoto_image',
            'img[src*="avatar"]',
            'div.EntityPhoto img',
        ]

        photo_url = None
        for selector in selectors:
            try:
                element = page.query_selector(selector)
                if element:
                    # Essayer d'obtenir l'URL
                    src = element.get_attribute('src')
                    if src and ('linkedin' in src or 'jpg' in src or 'png' in src):
                        photo_url = src
                        break

                    # Essayer data-src
                    data_src = element.get_attribute('data-src')
                    if data_src:
                        photo_url = data_src
                        break
            except:
                continue

        return photo_url

    except Exception as e:
        print(f"  ⚠️ Erreur: {str(e)[:50]}")
        return None

def scrape_photos(contacts):
    """Scrape les photos de profil de tous les contacts"""

    print("🔐 Connexion à LinkedIn...\n")

    with sync_playwright() as p:
        # Lancer le navigateur
        browser = p.chromium.launch(headless=False)  # headless=False pour voir ce qui se passe
        context = browser.new_context()
        page = context.new_page()

        # Connexion
        print("👤 Étape 1: Connexion à LinkedIn")
        print(f"   Email: {EMAIL}")
        print("   (Saisis ton mot de passe quand le navigateur s'ouvre)\n")

        page.goto("https://www.linkedin.com/login", wait_until="networkidle")

        # Attendre que l'utilisateur se connecte
        print("⏳ En attente de connexion manuelle...")
        print("   1. Saisis tes identifiants LinkedIn")
        print("   2. Fais la vérification 2FA si demandé")
        print("   3. Une fois connecté, le script continue automatiquement\n")

        # Attendre que l'utilisateur soit connecté (vérifier si on est redirecté)
        max_wait = 120  # 2 minutes
        start_time = time.time()
        while time.time() - start_time < max_wait:
            try:
                if "linkedin.com/feed" in page.url or "linkedin.com/in/" in page.url:
                    print("✓ Connexion réussie!\n")
                    break
            except:
                pass
            time.sleep(2)

        # Scraper les photos
        print(f"📸 Scraping de {len(contacts)} profils...\n")

        results = []
        success_count = 0

        for idx, contact in enumerate(contacts, 1):
            linkedin_url = contact.get('LinkedIn URL Profil', '')

            if not linkedin_url:
                print(f"[{idx}/{len(contacts)}] ⏭️  {contact['Prénom']} {contact['Nom']} - Pas d'URL LinkedIn")
                results.append({**contact, "Photo URL": None, "Photo URL (scraped)": datetime.now().isoformat()})
                continue

            print(f"[{idx}/{len(contacts)}] 🔗 {contact['Prénom']} {contact['Nom']}...", end=" ", flush=True)

            try:
                photo_url = extract_photo_url(page, linkedin_url)

                if photo_url:
                    print(f"✓ Photo trouvée")
                    results.append({**contact, "Photo URL": photo_url, "Photo URL (scraped)": datetime.now().isoformat()})
                    success_count += 1
                else:
                    print(f"⚠️  Photo non trouvée")
                    results.append({**contact, "Photo URL": None, "Photo URL (scraped)": datetime.now().isoformat()})

            except Exception as e:
                print(f"❌ Erreur: {str(e)[:30]}")
                results.append({**contact, "Photo URL": None, "Photo URL (scraped)": "error"})

            # Délai respectueux
            time.sleep(DELAY_BETWEEN_PROFILES)

        browser.close()

    # Sauvegarder les résultats
    print(f"\n\n✓ Scraping terminé!")
    print(f"   {success_count}/{len(contacts)} photos récupérées")
    print(f"\n💾 Sauvegarde dans {OUTPUT_FILE}...")

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"✓ Fichier sauvegardé: {OUTPUT_FILE}\n")

    return results

if __name__ == "__main__":
    print("="*60)
    print("🎯 LinkedIn Photo Scraper - Base de Recrutement Luxe")
    print("="*60)
    print()

    # Vérifier que le fichier Excel existe
    if not Path(EXCEL_FILE).exists():
        print(f"❌ Erreur: {EXCEL_FILE} non trouvé!")
        print(f"   Place ce script dans le même dossier que {EXCEL_FILE}")
        sys.exit(1)

    # Charger les contacts
    contacts = load_contacts()

    # Scraper les photos
    results = scrape_photos(contacts)

    print("="*60)
    print("✅ DONE!")
    print("="*60)
    print(f"\nFichier créé: {OUTPUT_FILE}")
    print("Prochaine étape: Ajouter les photos au HTML interactif")
